#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2017/8/12 21:52
# @Author  : GUO Ganggang
# @Site    : 
# @File    : obtain_data.py
# @Software: PyCharm

import codecs
import random
import io
import re
from itertools import islice
import openpyxl
from openpyxl.utils import get_column_letter
from os import listdir
import time
from gensim import utils
import gensim
import string
import os
import shutil
import numpy as np

import sys

reload(sys)
sys.setdefaultencoding('utf-8')

# 读入数据，并根据需要存储返回
def read_file(filePath):
    dedup_set = set()
    file_input_path = filePath + 'raw_data_from_DB\\eastmoney_guba_post_seg.csv'
    with codecs.open(file_input_path, 'rb', "utf-8") as infile:
        for line in infile:
            temp = line.strip().split('\t')
            if len(temp) == 1:
                dedup_set.add(temp[0])
            elif temp[1] == '':
                dedup_set.add(temp[0])
            else:
                continue
    print len(dedup_set)
    return dedup_set

# 东方财富股吧论坛数据，随机抽取6万条数据，并以每2500条为一个文件进行划分，分24批存入文件中
def divideDataSet_eastMoney(filePath):
    batch_num = {}
    temp_list = []
    flag = 400  # 上次设置600
    inter_flag = 0
    row_num = 0
    batch_i = 0

    # 读入已经存在的一批sentiment，并存入set中
    exit_set = set()
    exit_file_input_path = filePath + 'eastMoney_bbs_2\\eastMoney_bbs_batchs.csv'
    with codecs.open(exit_file_input_path, 'rb', "utf-8") as exit_infile:
        for exit in exit_infile:
            senti = exit.strip()
            if senti == '':
                continue
            else:
                exit_set.add(senti)
    print len(exit_set)

    # 随机抽取sentiment
    inFilePath = filePath + 'eastmoney_guba_post.csv'
    with codecs.open(inFilePath, "rb", "utf-8") as input_file:
        # for line in islice(input_file.readlines(), 1, None):
        for line in input_file:
            temp = line.strip()
            if temp == '':
                continue
            if temp in exit_set:
                continue
            row_num += 1
            interval_value = random.randint(466, 488) # 前次设置666, 688
            if row_num == flag:
                temp_list.append(temp)
                flag = flag + interval_value
                inter_flag += 1
                if (inter_flag == 2500) and (batch_i < 24):
                    batch_num[batch_i] = temp_list
                    batch_i += 1
                    inter_flag = 0
                    temp_list = []
    print len(batch_num)

    # 写入xlsx文件中
    for key in batch_num.keys():
        outputFilePath = filePath + 'eastMoney_bbs_2\\eastMoney_bbs_batch_' + str(key+1) + '.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        # ws = wb.create_sheet()
        ws.title = 'Sheet'
        row = 0
        for i in range(len(batch_num[key])):
            line = batch_num[key][i]
            temp = line.strip().split(',')
            if len(temp) < 2:
                continue
            content = ','.join(temp[1:])
            row += 1
            col = 1
            ws.cell(column=col, row=row, value=content) #.format(get_column_letter(col))
        wb.save(outputFilePath)

# 新浪财经新闻数据，随机抽取4万条数据，并以每2500条为一个文件进行划分，分16批存入文件中
def divideDataSet_sina(filePath):
    batch_num = {}
    temp_list = []
    flag = 45
    inter_flag = 0
    row_num = 0
    batch_i = 0
    inFilePath = filePath + 'sina_news.csv'
    with codecs.open(inFilePath, "rb", "utf-8") as input_file:
        # for line in islice(input_file.readlines(), 1, None):
        for line in input_file:
            temp = line.strip()
            if temp == '':
                continue
            row_num += 1
            interval_value = random.randint(40, 45)
            if row_num == flag:
                temp_list.append(temp)
                flag = flag + interval_value
                inter_flag += 1
                if (inter_flag == 2500) and (batch_i < 16):
                    batch_num[batch_i] = temp_list
                    batch_i += 1
                    inter_flag = 0
                    temp_list = []
    print len(batch_num)
    for key in batch_num.keys():
        outputFilePath = filePath + 'sinaFinance_news\\sinaFinance_news_batch_' + str(key + 1) + '.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        # ws = wb.create_sheet()
        ws.title = 'Sheet'
        row = 0
        # with codecs.open(filePath + 'sinaFinance_news\\sinaFinance_news_batch_' + str(key) + '.xlsx', 'w', encoding='utf-8') as output_file:
        for i in range(len(batch_num[key])):
            line = batch_num[key][i]
            # content = ','.join(line)
            row += 1
            col = 1
            # output_file.write(line + '\n')
            ws.cell(column = col,row = row,value = line.format(get_column_letter(col)))  #.format(get_column_letter(col))
        wb.save(outputFilePath)

# 读取xlsx内容,并写入文件中
def read_xlsx(filePath):
    # 获取文件夹中所有文件名称
    inFilePath = filePath + 'label_bbs_data\\labeled_first\\'
    fileNameFeature = '.xlsx'
    fileNameLabels = [f for f in listdir(inFilePath) if f.endswith(fileNameFeature)]
    print len(fileNameLabels)
    with open(filePath + 'label_bbs_data\\eastMoney_bbs_labeled_first.csv', 'w') as output_file:
        for fileName in fileNameLabels:
            print fileName
            #载入文件
            wb = openpyxl.load_workbook(inFilePath + fileName)
            #获取Sheet工作表
            ws = wb.get_sheet_by_name('Sheet')
            #按行读取
            for tuple_row in ws.rows:
                cell_row = []
                if len(tuple_row) < 2:
                    # print row[0]
                    continue
                for element_cell in tuple_row:
                    cell_row.append(str(element_cell.value))
                output_file.write(cell_row[1] + '\t' + cell_row[0] + '\n')

# 读入文件，去掉重复内容,再写入文件
def dedup(filePath):
    uid_screenName_score = {}
    infilePath =filePath +  'caijing_weibo_uid_info_filter_0913.csv'
    outfilePath =filePath +  'caijing_weibo_uid_info_filter_0913_dedup.csv'

    with codecs.open(infilePath, "rb", "utf-8") as input_file:
        for line in input_file.readlines():
            temp = line.strip().split(',')
            # stock_name = re.sub('\*', '', temp)
            content = ','.join(temp[1:])
            # if temp[0] not in uid_screenName_score.keys():
            uid_screenName_score[temp[0]] = content
            # else:
            #     continue
    print len(uid_screenName_score)

    # 排序
    UIDscreenName_score = {}
    for key in uid_screenName_score.keys():
        screenName_score = uid_screenName_score[key].strip().split(',')
        new_key = str(key) + ',' + str(screenName_score[0])
        UIDscreenName_score[new_key] = float(screenName_score[1])

    socre_sort = sorted(UIDscreenName_score.iteritems(), key=lambda d: d[1], reverse=True)

    with open(outfilePath, 'w') as output_file:
        for i in range(len(socre_sort)):
            output_file.write(socre_sort[i][0] + ',' + str(socre_sort[i][1]) + '\n')

# 过滤分词后的股吧数据,过滤掉空、全英文和词（字）数少于3的句子
def obtain_special_sentiments(filePath):
    infilePath = filePath + 'label_bbs_data\\eastMoney_bbs_labeled_total_seg.csv'
    outfilePath = filePath + 'label_bbs_data\\eastMoney_bbs_labeled_total_seg_w2v.csv'
    # infilePath = filePath + 'raw_data_from_DB\\eastmoney_guba_post_seg.csv'
    # outfilePath = filePath + 'raw_data_from_DB\\eastmoney_guba_post_seg_w2v.csv'
    pattern = re.compile('\s+')
    # punc = '''！？｡＂＃＄％＆＇（）＊＋，－／：；＜＝＞＠［＼］＾＿｀｛｜｝～｟｠｢｣､、〃》「」『』【】〔〕〖〗〘〙〚〛〜〝〞〟〰〾〿–—‘’‛“”„‟…‧﹏.'''
    with open(outfilePath, 'w') as output_file:
        with codecs.open(infilePath, "rb", "utf-8") as input_file:
            for line in input_file:
                temp = line.strip().split('\t')
                if (len(temp) != 2):
                    continue
                content = ' '.join(temp[1:]).strip()
                # if len(content.split(' ')) < 3:
                #     continue
                text = ''.join(temp[1:])
                clear_eng_pun = str(text).translate(None, string.punctuation)  # clear 英文标点
                clear_chi_pun = re.sub(ur"[%s]+" % zhon.hanzi.punctuation, "", clear_eng_pun.decode("utf-8")) # clear 中文标点
                clear_tab = re.sub(pattern, '', clear_chi_pun)
                unicode_cov_str = clear_tab.encode("utf-8").strip()
                # if (unicode_cov_str == ''):
                #     continue
                if (unicode_cov_str.isalpha() == True):
                    if unicode_cov_str == 'sb' or unicode_cov_str == 'SB' or unicode_cov_str == 'TMD' or unicode_cov_str == 'tmd':
                        output_file.write(temp[0] + '\t' + content + '\n')
                    else:
                        continue
                else:
                    output_file.write(temp[0] + '\t' + content + '\n')

# 读入文件夹中所有文件，筛选后，写入另外文件中
def read_filtering_write(filePath):
    inFilePath = filePath + 'match_title_only_exectly\\'
    outFilePath = filePath + 'match_title_only_exectly.csv'
    fileNameFeature = '.csv'
    fileNameLabels = [f for f in listdir(inFilePath) if f.endswith(fileNameFeature)]
    with open(outFilePath, 'w') as output_file:
        for k in range(len(fileNameLabels)):
            nameFile = ''.join(fileNameLabels[k])
            # print chardet.detect(nameFile)
            stock_name = nameFile.decode('gbk').replace('.csv', '')
            # stock_name_utf =  repr(stock_name) #.strip().split('\\')
            # stock_name_str = '\\'.join(stock_name_utf)
            filter_sum = 0
            with codecs.open(inFilePath + nameFile, "rb", "utf-8") as input_file:
                for line in input_file.readlines():
                    obtain_list = []
                    temp = line.strip().split(',')
                    if len(temp) < 7:
                        print len(temp),line.strip()
                        continue
                    news_title = ','.join(temp[0:(len(temp) - 6)])
                    # print stock_name_str,news_title
                    if stock_name in news_title:
                        obtain_list.append(stock_name)
                        obtain_list.append(temp[len(temp) - 6])
                        obtain_list.append(temp[len(temp) - 1])
                        obtain_list.append(temp[len(temp) - 2])
                        obtain_list.append(temp[len(temp) - 3])
                        obtain_list.append(temp[len(temp) - 4])
                        obtain_list.append(news_title)
                        filter_sum += 1
                    else:
                        continue
                    line_new = '\t'.join(obtain_list)
                    output_file.write(line_new + '\n')
            # print stock_name,str(filter_sum)

# 按年、新闻来源进行统计
def static_by_year_source(ipath):
    # 需要统计的新闻数据源
    news_main_sources = [u'中国证券报',u'证券日报',u'证券时报',u'中国经营报',u'上海证券报',u'经济观察报',u'21世纪经济报']

    # 获取股票名称
    stocks_names = set()
    inFilePath = ipath + 'sina_news_static.csv'
    with codecs.open(inFilePath, "rb", "utf-8") as inputfile:
        for line in islice(inputfile.readlines(), 0, None):
            temp = line.strip().split(',')
            if len(temp) != 3:
                print line.strip()
                continue
            clean_stock_name = re.sub('"','',temp[2])
            stocks_names.add(clean_stock_name)
    print len(stocks_names)

    # 实现针对每支股票按年和新闻来源统计
    stockName_static_by_year = {}
    for every_stock in stocks_names:
        static_by_year_sources = {}
        with codecs.open(inFilePath, "rb", "utf-8") as inputfile:
            # for line in islice(inputfile.readlines(), 0, None):
            for line in inputfile:
                temp = line.strip().split(',')
                if len(temp) != 3:
                    # print line.strip()
                    continue
                clean_stock_name = re.sub('"', '', temp[2])
                clean_post_date = re.sub('"', '', temp[1])
                clean_post_year = clean_post_date[0:4]
                clean_news_source = re.sub('"', '', temp[0])
                clean_news_source = re.sub('\《', '', clean_news_source)
                clean_news_source = re.sub('\》', '', clean_news_source)
                if (clean_stock_name == every_stock) and (clean_news_source in news_main_sources):
                    static_by_year_sources[clean_post_year] = static_by_year_sources.get(clean_post_year,0) + 1
        stockName_static_by_year[every_stock] = static_by_year_sources
        print every_stock,static_by_year_sources
    print len(stockName_static_by_year)

    # 将统计结果写入文件
    years = ['2010','2011','2012','2013','2014','2015','2016','2017']
    years_str = ','.join(years)
    outFilePath = ipath + 'stock_statistics_results_by_year_sources.csv'
    with open(outFilePath, 'w') as output_file:
        output_file.write('stock_name' + ',' + years_str + '\n')
        for outer_key in stockName_static_by_year.keys():
            static_results_by_year = []
            for i in range(len(years)):
                if years[i] in stockName_static_by_year[outer_key].keys():
                    static_results_by_year.append(str(stockName_static_by_year[outer_key][years[i]]))
                else:
                    static_results_by_year.append(str(0))
            static_results_by_year_str = ','.join(static_results_by_year)
            output_file.write(outer_key + ',' + static_results_by_year_str + '\n')

# 匹配量文件中的数据,并做统计，生成字典
def match_files_static(filePath):
    # inFilePath_fir = filePath + 'custom_dic\\stock_name_quixng.csv'
    # infilePath_sec = filePath + 'raw_data_from_DB\\eastmoney_guba_post.csv'
    # outfilePath = filePath + 'stock_index_name.dic'

    inFilePath_fir = filePath + 'gainiangu_description_tfidf_Nwords_makeup_extend_by_w2v_seg_dic.csv'
    infilePath_sec = 'D:\\SMU_WORK\\event_gainiangu_building_model\\weibo\\stockcomment.txt'
    outfilePath = filePath + 'gainiangu_description.dic'

    word_static = {}
    dedup_word = set()
    with codecs.open(inFilePath_fir, "rb", "utf-8") as inputfile:
        for line in islice(inputfile.readlines(), 0, None):
            temp = line.strip() #.split('	')
            # if len(temp) != 2:
            #     print line.strip()
            #     continue
            if temp not in dedup_word:
                word_static[temp] = 0
                dedup_word.add(temp)
        print len(word_static)

    # with codecs.open(infilePath_sec, "rb", "utf-8") as input_file:
    #     # for line in islice(input_file.readlines(), 0, None):
    #     for line in input_file:
    #         # temp = line.strip().split(',')
    #         # if len(temp) < 2:
    #         #     print line.strip()
    #         #     continue
    #         temp = line.strip()
    #         # stock_name = re.sub('\*', '', temp[0])
    #         for key in dedup_word:
    #             if key in temp:
    #                 word_static[key] = word_static.get(key, 0) + 1

    contents = []
    with codecs.open(infilePath_sec, "rb", "utf-8") as input_file:
        # for line in islice(input_file.readlines(), 0, None):
        for line in input_file:
            temp = line.strip().split('\' \'')
            if len(temp) < 4:
                continue
            contents.append(temp[3])
    print len(contents)
    flag = 0
    for key in dedup_word:
        print key,str(flag+1)
        for content in contents:
            if key in content:
                word_static[key] = word_static[key] + 1
    words_sort = sorted(word_static.iteritems(), key=lambda d: d[1], reverse=True)

    with open(outfilePath, 'w') as output_file:
        for i in range(len(words_sort)):
            output_file.write(str(i+1) + '\t' + words_sort[i][0] + '\t'+ str(words_sort[i][1]) + '\n')

# 匹配两文件中的数据
def match_files(filePath):
    inFilePath_fir = filePath + 'label_bbs_data\\eastMoney_bbs_labeled_first.csv'
    infilePath_sec = filePath + 'label_bbs_data\\eastMoney_bbs_labeled_second.csv'
    outfilePath_fir = filePath + 'label_bbs_data\\eastMoney_bbs_labeled_not_matched_neutral.csv'
    outfilePath_sec = filePath + 'label_bbs_data\\eastMoney_bbs_labeled_matched.csv'
    outfilePath_thr = filePath + 'label_bbs_data\\eastMoney_bbs_labeled_not_matched_paradox.csv'

    label_text_fir = {}
    fir_set = set()
    with codecs.open(inFilePath_fir, "rb", "utf-8") as inputfile:
        # for line in islice(inputfile.readlines(), 0, None):
        for line in inputfile:
            temp = line.strip().split('	')
            if len(temp) < 2:
                continue
            text = '	'.join(temp[1:])
            if text not in fir_set:
                label_text_fir[text] = temp[0]
                fir_set.add(text)
    print len(label_text_fir)


    label_text_sec = {}
    label_text_thr = {}
    label_text_zero = {}
    with codecs.open(infilePath_sec, "rb", "utf-8") as input_file:
        # for line in islice(input_file.readlines(), 0, None):
        content_special = [u'创建了投资组合',u'创建了模拟组合',u'创建了实盘组合']
        for line in input_file:
            temp = line.strip().split('	')
            if len(temp) < 2:
                continue
            text = '	'.join(temp[1:])
            # for con in content_special:
            if content_special[0] in text:
                label_text_sec[text] = '0'
                continue
            if content_special[1] in text:
                label_text_sec[text] = '0'
                continue
            if text in fir_set:
                if temp[0] == label_text_fir[text]:
                    label_text_sec[text] = temp[0]
                elif str(temp[0]) == '0' and str(label_text_fir[text]) == '-1':
                    label_text_sec[text] = label_text_fir[text]
                elif str(temp[0]) == '0' and str(label_text_fir[text]) == '1':
                    label_text_zero[text] = temp[0]
                else:
                    label_text_thr[text] = temp[0]
            else:
                label_text_thr[text] = temp[0]
    print len(label_text_zero),len(label_text_sec),len(label_text_thr)

    with open(outfilePath_fir, 'w') as output_file:
        for key in label_text_zero.keys():
            output_file.write(str(label_text_zero[key]) + '\t' + key + '\n')
    with open(outfilePath_sec, 'w') as output_file:
        for key in label_text_sec.keys():
            output_file.write(str(label_text_sec[key]) + '\t' + key + '\n')
    with open(outfilePath_thr, 'w') as output_file:
        for key in label_text_thr.keys():
            output_file.write(str(label_text_thr[key]) + '\t' + key + '\n')

# 判断文件夹是否存在，不存在就自动生成
def createIfNotExists(path):
    if os.path.exists(path):
        print "path already exists!"
    else:
        os.mkdir(path)
        print "dir created"

# 迭代获取文件夹中全部文件路径与名称，并存入list中返回
def listdirfiles(filePath):
    folderNameFeature = 'out_josn_files'
    files_path_name = []
    for folder in os.listdir(filePath):
        folder_path = os.path.join(filePath, folder)
        # print folder_path
        if folder.startswith(folderNameFeature) and os.path.isdir(folder_path):
            for file_name in os.listdir(folder_path):
                file_path = os.path.join(folder_path, file_name)
                files_path_name.append(file_path)
    print len(files_path_name)
    return files_path_name

# python之文件操作-复制、剪切、删除等
def movie_files(currentFilePath,destinationFilePath):
    files_path_name = listdirfiles(currentFilePath)
    # os.rename("path/to/current/file.foo", "path/to/new/desination/for/file.foo")
    createIfNotExists(destinationFilePath)
    for i in range(len(files_path_name)):
        file_name = os.path.split(files_path_name[i])[1]
        destination_path_file = os.path.join(destinationFilePath, file_name)
        shutil.move(files_path_name[i], destination_path_file)

# 获取满足股龄的股票代码
def obtain_stockcode_by_guling(filePath):

    inFilePath_fir = filePath + 'news_data_from_youkuang\\hushen_all_stocks_20070101_before.csv'
    infilePath_sec = filePath + 'news_data_from_youkuang\\hushen_all_stocks_20170331_before.csv'
    outfilePath = filePath + 'news_data_from_youkuang\\hushen_stockcode_by_guling.csv'

    stock_2007 = []
    with codecs.open(inFilePath_fir, "rb", "utf-8") as inputfile:
        for line in islice(inputfile.readlines(), 0, None):
            temp = line.strip().split('	')
            if len(temp) < 2:
                print line.strip()
                continue
            stock_2007.append(temp[0])
    print len(stock_2007)

    selected_stockcode = []
    with codecs.open(infilePath_sec, "rb", "utf-8") as input_file:
        # for line in islice(input_file.readlines(), 0, None):
        for line in input_file:
            temp = line.strip().split('	')
            if len(temp) < 2:
                print line.strip()
                continue
            if temp[0] in stock_2007:
                selected_stockcode.append(temp[0])
    print len(selected_stockcode)

    with open(outfilePath, 'w') as output_file:
        for stockcode in selected_stockcode:
            output_file.write(str(stockcode) + '\n')

# 清理从数据库导出来的数据，将由于换行引起的错误消除
def adjust_dataset(filePath):
    infilePath  = filePath + 'raw_data_from_DB\\eastmoney_guba_post.csv'
    outfilePath_1 = filePath + 'raw_data_from_DB\\eastmoney_guba_post_new.csv'
    outfilePath_2 = filePath + 'raw_data_from_DB\\eastmoney_guba_post_new_new.csv'

    # 消除换行符
    with open(outfilePath_1, 'w') as output_file:
        with codecs.open(infilePath, "rb", "utf-8") as input_file:
            previous_id_text = ''
            for line in input_file:
                temp = line.strip().split(',')
                if line.strip() == '':
                    continue
                else:
                    if str(temp[0]).isdigit() and len(temp[0]) == 9:
                        previous_id_text = ''
                        output_file.write(line.strip() + '\n')
                        previous_id_text = line.strip()
                    else:
                        previous_id_text = previous_id_text + line.strip()
                        output_file.write(previous_id_text + '\n')


    # 消除上一步产生的重复行
    with open(outfilePath_2, 'w') as output_file:
        with codecs.open(outfilePath_1, "rb", "utf-8") as input_file:
            previous_id = ''
            previous_text = ''
            for line in input_file:
                temp = line.strip().split(',')
                if len(temp) < 2:
                    print line.strip()
                    continue
                else:
                    if previous_id == str(temp[0]) or previous_id == '':
                        previous_id = str(temp[0])
                        previous_text = ','.join(temp[1:])
                    else:
                        output_file.write(previous_id + ',' + previous_text + '\n')
                        previous_id = str(temp[0])
                        previous_text = ','.join(temp[1:])

# 过滤文件
def filter_files(filePath):
    inFilePath_fir = filePath + 'news_data_from_youkuang\\filter_stock_by_turnover_guling.csv'
    infilePath_sec = filePath + 'news_data_from_youkuang\\eastmoney_guba_post_barname.csv'
    outfilePath = filePath + 'news_data_from_youkuang\\filter_stock_by_turnover_guling_comments.csv'

    stock_name_20170331 = []
    with codecs.open(inFilePath_fir, "rb", "utf-8") as inputfile:
        for line in islice(inputfile.readlines(), 0, None):
            temp = line.strip()
            # if len(temp) != 2:
            #     print line.strip()
            #     continue
            # stock_name = re.sub('\*ST', '', temp[1])
            stock_name_20170331.append(temp)

    print len(stock_name_20170331)

    stock_comments_num_statistic = {}
    with codecs.open(infilePath_sec, "rb", "utf-8") as input_file:
        # for line in islice(input_file.readlines(), 0, None):
        for line in input_file:
            temp = line.strip().split(',')
            stock_name = re.sub(u'吧', '', temp[1])
            if stock_name in stock_name_20170331:
                stock_comments_num_statistic[stock_name] = stock_comments_num_statistic.get(stock_name, 0) + 1
            else:
                continue
    print len(stock_comments_num_statistic)
    stock_comments_num_sort = sorted(stock_comments_num_statistic.iteritems(), key=lambda d: d[1], reverse=True)

    with open(outfilePath, 'w') as output_file:
        for i in range(len(stock_comments_num_sort)):
            output_file.write(stock_comments_num_sort[i][0] + '\t' + str(stock_comments_num_sort[i][1]) + '\n')

def filter_file(filePath):
    # infilePath = filePath + 'label_bbs_data\\eastMoney_bbs_labeled_not_matched_paradox.csv'
    # outfilePath_fir = filePath + 'label_bbs_data\\eastMoney_bbs_labeled_not_matched_paradox_01.csv'
    # outfilePath_sec = filePath + 'label_bbs_data\\eastMoney_bbs_labeled_not_matched_paradox_02.csv'
    # label_text_one = {}
    # label_text_other = {}
    # texts = set()
    # with codecs.open(infilePath, "rb", "utf-8") as input_file:
    #     for line in input_file:
    #         temp = line.strip().split('	')
    #         text = '	'.join(temp[1:])
    #         if text in texts:
    #             continue
    #         else:
    #             texts.add(text)
    #         if str(temp[0]) == '1':
    #             label_text_one[text] = temp[0]
    #         else:
    #             label_text_other[text] = temp[0]
    #
    # with open(outfilePath_fir, 'w') as output_file:
    #     for key in label_text_one.keys():
    #         output_file.write(str(label_text_one[key]) + '\t' + key + '\n')
    # with open(outfilePath_sec, 'w') as output_file:
    #     for key in label_text_other.keys():
    #         output_file.write(str(label_text_other[key]) + '\t' + key + '\n')

    infilePath = filePath + 'gainiangu_description_tfidf_Nwords_makeup_extend_by_w2v_20w_clean_tfidf.csv'
    # outfilePath = filePath + 'label_bbs_data\\pos_data_set.csv'
    # finance_keywords = ['股','金融','财经','经济']
    outfilePath = filePath + 'gainiangu_description_tfidf_Nwords_makeup_extend_by_w2v_20w_clean_tfidf_41words.csv'
    with open(outfilePath, 'w') as output_file:
        with codecs.open(infilePath, "rb", "utf-8") as input_file:
            for line in input_file:
                temp = line.strip().split('\t')
                # print len(temp)
                # print line.strip()
                words = temp[1].strip().split(' ')
                # print len(words)
                word_tfidf = {}
                words_sort = []
                for word in words:
                    # print word
                    key_value = word.strip().split(',')
                    # print key_value[0]
                    word_tfidf[key_value[0]] = key_value[1]
                words_tfidf_sort = sorted(word_tfidf.iteritems(), key=lambda d: d[1], reverse=True)
                if len(words_tfidf_sort) > 41:
                    for k in range(41):
                        # print words_tfidf_sort[k]
                        words_sort.append(words_tfidf_sort[k][0])
                else:
                    for j in range(len(words_tfidf_sort)):
                        words_sort.append(words_tfidf_sort[j][0])
                # print words_sort
                words_sort_str = ','.join(words_sort)

                # temp = line.strip().split(r'","')
                # if str(temp[0]) == '0':
                # end_num = len(temp) - 3
                # text = ''.join(temp[3:end_num])
                # if text == '':
                #     continue
                output_file.write(temp[0] + ',' + words_sort_str + '\n')
                # else:
                #     continue
                    # output_file.write(line.strip() + '\n')

def match_uid_screeen_name(filePath):
    inFilePath_fir = filePath + 'finance_weibo_uid_info_score.csv'
    infilePath_sec = filePath + 'caijing_weibo_uid.csv'
    outfilePath = filePath + 'caijing_weibo_uid_info.csv'

    uid_screenName_score = {}
    with codecs.open(inFilePath_fir, "rb", "utf-8") as inputfile:
        for line in islice(inputfile.readlines(), 0, None):
            temp = line.strip().split(',')
            if len(temp) < 3:
                continue
            content = ','.join(temp[1:])
            uid_screenName_score[temp[0]] = content
    print len(uid_screenName_score)

    matched_uid_screenName_score = {}
    with codecs.open(infilePath_sec, "rb", "utf-8") as input_file:
        for line in input_file:
            temp = line.strip()
            if temp in uid_screenName_score.keys():
                matched_uid_screenName_score[temp] = uid_screenName_score[temp]
    print len(matched_uid_screenName_score)

    with open(outfilePath, 'w') as output_file:
        for key in matched_uid_screenName_score.keys():
            output_file.write(str(key) + ',' + matched_uid_screenName_score[key] + '\n')

def match_much_files(filePath):

    inFilePath = filePath + 'individual_stock_discription\\'
    outfilePath = filePath + 'individual_stock_description.csv'
    fileNameLabels = [f for f in listdir(inFilePath) if f.endswith('_seg.csv')]

    individual_stock_description = {}
    print len(fileNameLabels)

    for fileName in fileNameLabels:
        with codecs.open(inFilePath + fileName, "rb", "utf-8") as inputfile:
            for line in inputfile:
                temp = line.strip().split('\t')
                if len(temp) < 2:
                    print fileName,line.strip()
                    continue
                # content = ','.join(temp[1:])
                if temp[0] in individual_stock_description.keys():
                    individual_stock_description[temp[0]] =individual_stock_description[temp[0]] + ' '+ temp[1]
                else:
                    individual_stock_description[temp[0]] = temp[1]
    print len(individual_stock_description)

    with open(outfilePath, 'w') as output_file:
        for key in individual_stock_description.keys():
            output_file.write(str(key) + '\t' + individual_stock_description[key] + '\n')

def clean_seg_sentiments(filePath):
    infilePath = filePath + 'eastmoney_guba_post_urls_completed_parsed_seg.csv'
    outfilePath = filePath + 'eastmoney_guba_post_urls_completed_parsed_seg_w2v.csv'

    clean_list = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '0', "'''", "''", '．', '……', ',', '﹍', '▁', '▂▃', \
                  '＂', '／', '〜', '•', '➊', '➋', '➌', '➍', '➎', '➏', '﹑', '´', '╬╬', '＼', '﹎', '﹏', '▃▃', '▆▇', '◥◣', \
                  '︻', '.', '%', '﹌', 'ミ', '∟', '☆', '…', '➠', '➕', '•', 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', '▃', '█',  \
                  '■■',']', 'мī','ч','html','csv', u'空',u'版', u'网友',u'网页' ,\
                  'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'ゞ', '∷', '▌', \
                  '◢◤', '▆',u'博客', \
                  '￡', '╬', 'з', 'ヽ', '｀', '︱', '┢┦', '※', '�', '▄', '┻═', '┳一', '－', '≦', '〆', '▌▌', '▅▆', '▇◤', '▆▆',\
                  '◥', '■■', '▋', '◥◤', '■■']
    values = set()
    with codecs.open(outfilePath, "w", "utf-8") as output_file:
        with codecs.open(infilePath, "rb", "utf-8") as inputStockCode:
            for line in inputStockCode:
                if line != '':
                    temp = line.strip().split(' ')
                    clean_temp = []
                    for i in range(len(temp)):
                        flag = '0'
                        for string in clean_list:
                            if string in temp[i]:
                                flag = '1'
                                break
                        if flag == '1':
                            continue
                        else:
                            clean_temp.append(temp[i])
                    lenth = len(clean_temp)
                    vec = " ".join(clean_temp[0:])
                    if lenth >= 3:
                        if vec in values:
                            continue
                        else:
                            output_file.write(vec + '\n')
                            if len(values) <= 10:
                                values.add(vec)
                            else:
                                values.clear()
                        # output_file.write(vec + '\n')  # + '\t' + temp[2] + '\t' + temp[3]
                        # elif lenth < k:
                        #     print lenth,k
                        #     print vec
                else:
                    continue

def one_hot_embedding(filePath):
    inFilePath = filePath + 'gainiangu_gupiaochi_relation.csv'
    outfilePath = filePath + 'gainiangu_gupiaochi_relation_embedding.csv'
    gainianName_stocks = {}
    stocks_list = []
    with codecs.open(inFilePath, "rb", "utf-8") as inputfile:
        for line in inputfile:
            temp = line.strip().split('\t')
            if len(temp) < 2:
                continue
            stocks = temp[1].strip().split(',')
            for stock in stocks:
                if stock not in stocks_list:
                    stocks_list.append(stock)
            gainianName_stocks[temp[0]] = temp[1]

    print len(gainianName_stocks),len(stocks_list)  #405 3283

    with open(outfilePath, 'w') as output_file:
        title = ','.join(stocks_list)
        output_file.write(u'概念名称' + '\t' + title + '\n')
        for key in gainianName_stocks.keys():
            gs_embedding = []
            for k in range(len(stocks_list)):
                gs_embedding.append(0.0)
            gainian_stocks = gainianName_stocks[key].strip().split(',')
            for j in range(len(stocks_list)):
                if stocks_list[j] not in gainian_stocks:
                    continue
                else:
                    gs_embedding[j] = 1.0 / len(gainian_stocks)
            values = ','.join([str(value) for value in gs_embedding])
            output_file.write(str(key) + '\t' + values + '\n')

def matrix_operation(filePath):
    fileNames = ['gainiangu_description_tfidf_Nwords_makeup_extend_by_w2v_extend2_remove_tfidf_20words_embedding.csv','gainiangu_description_tfidf_Nwords_makeup_extend_by_w2v_extend2_remove_tfidf_20words_dict_embedding.csv']
    gainian_dirc = {}
    seed_words_dirc = {}
    dicts = [gainian_dirc,seed_words_dirc]

    for i in range(len(fileNames)):
        inFilePath = filePath + fileNames[i]
        with codecs.open(inFilePath, "rb", "utf-8") as inputfile:
            for line in inputfile:
                temp = line.strip().split('\t')
                if len(temp) < 2:
                    print line.strip()
                    continue
                vec = temp[1].strip().split(',')
                dicts[i][temp[0]] = vec
    a = len(seed_words_dirc)
    b = len(gainian_dirc)
    outfilePath = filePath + 'seed_words_gainian_distance_%s' %a +'_%s' %b +  '.csv'
    with open(outfilePath, 'w') as output_file:
        row_titles = ','.join(gainian_dirc.keys())
        output_file.write(u'横纵标题' + ',' + row_titles + '\n')
        for seed_word in seed_words_dirc.keys():
            values = []
            for gainian in gainian_dirc.keys():
                vec1 = np.array(seed_words_dirc[seed_word], dtype='float64')
                vec2 = np.array(gainian_dirc[gainian], dtype='float64')
                #dist = np.linalg.norm(vec1 - vec2)  #效果一样
                dist = np.sqrt(np.sum(np.square(vec1 - vec2)))
                values.append(str(dist))
            values_str = ','.join(values)
            output_file.write(seed_word + ','+ values_str + '\n')
            # output_file.write(values_str + '\n')

def identit_conv(filePath):
    fileNames = ['seed_words_gainian_distance_7591_393.csv',
                 'event_hot_keywords_dict_embedding_7591_128.csv']
    seed_words_gainian_distance_dirc = {}
    seed_words_dirc = {}
    dicts = [seed_words_gainian_distance_dirc, seed_words_dirc]

    for i in range(len(fileNames)):
        inFilePath = filePath + fileNames[i]
        with codecs.open(inFilePath, "rb", "utf-8") as inputfile:
            for line in inputfile:
                temp = line.strip().split(',')
                if len(temp) < 2:
                    print line.strip()
                    continue
                dicts[i][temp[0]] = ','.join(temp[1:])
    del seed_words_gainian_distance_dirc[u'横纵标题']
    a = len(seed_words_dirc)
    b = len(seed_words_gainian_distance_dirc)
    print a,b

    outfilePath_fir = filePath +  'seed_words_gainian_distance_7591_393_noTitles.csv'
    outfilePath_sec = filePath + 'event_hot_keywords_dict_embedding_7591_128_noTitles.csv'
    with open(outfilePath_sec, 'w') as output_file:
        for key in seed_words_dirc.keys():
            output_file.write(seed_words_dirc[key] + '\n')
    with open(outfilePath_fir, 'w') as output_file:
        for key in seed_words_dirc.keys():
            output_file.write(seed_words_gainian_distance_dirc[key] + '\n')


if __name__ == '__main__':
    filePath = 'D:\\east_money_guba\\'
    # divideDataSet_eastMoney(filePath)
    # divideDataSet_sina(filePath)
    # dedup(filePath)
    # read_xlsx(filePath)
    # obtain_special_sentiments(filePath)
    # filePath = 'D:\\data\\paper_data\\data_collection\\bjinfobank\\'
    # read_filtering_write(filePath)
    # filePath = 'D:\\data\\paper_data\\'
    # filter_file(filePath)
    # static_by_year_source(filePath)
    # match_files(filePath)
    # folders_list = ['clawer_guba_7','clawer_guba_8','clawer_guba_9','clawer_guba_13']
    # destinationFilePath_list = ['out_josn_files_22', 'out_josn_files_23', 'out_josn_files_24', 'out_josn_files_25']
    # for k in range(4):
    #     currentFilePath = 'D:\\SMU_WORK\\' + folders_list[k]
    #     destinationFilePath = 'E:\\eastmoney_guba\\' + destinationFilePath_list[k]
    #     movie_files(currentFilePath,destinationFilePath)
    # obtain_special_sentiments(filePath)
    # obtain_stockcode_by_guling(filePath)
    # adjust_dataset(filePath)
    # filter_files(filePath)
    # filter_files(filePath)
    # filter_file(filePath)
    # match_uid_screeen_name(filePath)

    # filePath = 'D:\\SMU_WORK\\event_gainiangu_building_model\\weibo\\'
    # filter_file(filePath)
    # clean_seg_sentiments(filePath)

    # filePath = 'C:\\Users\\lenovo\\Desktop\\intelligence_concept_stock\\events_gainians_stocks_models_dataset\\words_gainian_model_dataset\\v3.0\\'
    # filter_file(filePath)
    # match_much_files(filePath)
    # one_hot_embedding(filePath)

    # matrix_operation(filePath)
    # identit_conv(filePath)

    # identit_conv(filePath)

    # match_files_static(filePath)

    clean_seg_sentiments(filePath)

